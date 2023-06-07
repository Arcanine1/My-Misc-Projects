
import requests
import numpy as np #Used to solve overdetermined simultaneous equations where negative values are fine
from scipy.optimize import nnls #Used to solve overdetermined simultaneous equations where negative values are not fine
from openpyxl import load_workbook #Used to export data to excel sheet

n=1
just_teams_array=[]
just_scores_array=[]
expected_score_array=[]

while(n<33):

    print(n)

    URL="https://www.thebluealliance.com/match/2021mibg_qm"+str(n)
    page = requests.get(URL)
    code = str(page.text)  # These two lines get the HTML and put it into a string so that it can be searched
    i = 1
    index = 0
    while i < 7:
        index = code.find("data-team", index + 1)
        start = index + 14
        end = index + 18
        Team = code[start:end]
        numeric_filter = filter(str.isdigit, Team)
        Team = "".join(numeric_filter)
        just_teams_array.append(int(Team))
        i = i + 1

    # Gets Red Score from TBA
    index = code.find('dScore"')
    index = code.find('span',index)
    start = index + 5
    end = index + 25
    Score = code[start:end]
    print(Score)
    numeric_filter = filter(str.isdigit, Score)
    Score = "".join(numeric_filter)
    just_scores_array.append(int(Score))

    # Gets Blue Score from TBA
    index = code.find('eScore"')
    index = code.find('span', index)
    start = index + 5
    end = index + 25
    Score = code[start:end]
    numeric_filter = filter(str.isdigit, Score)
    Score = "".join(numeric_filter)
    just_scores_array.append(int(Score))

    n=n+1


#Is the path of the excel sheet where the data will go
path='/Users/ayushgarg/Desktop/Web Scraper Test.xlsx'
workbook = load_workbook(path)
ws = workbook.get_sheet_by_name("Sheet1")

lenght_score=len(just_scores_array)
numberofmatch=int(lenght_score/2)



b=0
a=1
c=0
while(2>1):
    if(ws.cell(row=a,column=1).value==None):
        break;
    a=a+1
while(c<numberofmatch*2):
    ws.cell(row=a, column=5).value = just_teams_array[b]
    ws.cell(row=a, column=6).value = just_teams_array[b+1]
    ws.cell(row=a, column=7).value = just_teams_array[b+2]
    ws.cell(row=a, column=8).value = just_teams_array[b+3]
    ws.cell(row=a, column=9).value = just_teams_array[b+4]
    ws.cell(row=a, column=10).value = just_teams_array[b+5]
    ws.cell(row=a, column=13).value = just_scores_array[c]
    ws.cell(row=a+1, column=5).value = just_teams_array[b+4]
    ws.cell(row=a+1 ,column=6).value = just_teams_array[b+5]
    ws.cell(row=a+1, column=7).value = just_teams_array[b+3]
    ws.cell(row=a+1, column=8).value = just_teams_array[b+1]
    ws.cell(row=a+1, column=9).value = just_teams_array[b+2]
    ws.cell(row=a+1, column=10).value = just_teams_array[b]
    ws.cell(row=a+1, column=13).value = just_scores_array[c+1]
    a = a + 2
    c = c + 2
    b=b+6

lenght_score=len(just_scores_array)
team_numberer=set(just_teams_array) #Gets rid of duplicates
team_numberer=sorted(team_numberer) #Makes it go in ascending order

numberofteam=int(len(team_numberer))
numberofmatch=int(lenght_score/2)

#Creates a 2-D array which has the number of teams as one dimension and the number of scores as the other
Team = [[0 for k in range(numberofteam)] for j in range(numberofmatch*2)]

backup=just_teams_array
e=1
d=0
while(e-1<int(len(team_numberer))):
    while(int(len(just_teams_array))>d):
        if(team_numberer[e-1]==just_teams_array[d]):
            just_teams_array[d]=e
        d=d+1
    e=e+1
    d=0


a=0
b=0

while(a<numberofmatch*2):
    print(just_teams_array)
    Team[a][just_teams_array[b]-1]=1
    b=b+1
    Team[a][just_teams_array[b]-1]=1
    b=b+1
    Team[a][just_teams_array[b]-1]=1
    b=b+1
    a=a+1

A=np.array(Team)
B=np.array(just_scores_array)
Scores= nnls(A, B)[0]

y=0
v=0
expected_score=0
while(y<numberofmatch*2):
    while(v<numberofteam):
        if(Team[y][v]==1):
            expected_score=expected_score+Scores[v]
        v=v+1
    y=y+1
    v=0
    expected_score_array.append(expected_score)
    expected_score=0


defense_team_array=Team
#Switches up the arrays so that the opposing alliance array is in the slot where the other alliance was
j=0
while(j<numberofmatch*2):
    defense_team_array[j],defense_team_array[j+1]=defense_team_array[j+1],defense_team_array[j]
    j=j+2

A=np.array(defense_team_array)
B=np.array(np.array(expected_score_array)-np.array(np.array(just_scores_array)))
Defense= np.linalg.lstsq(A, B,rcond=None)[0]



#Finds the first unoccupied row of the excel sheet
a=1
while(2>1):
    if(ws.cell(row=a,column=1).value==None):
        break;
    a=a+1
b=0

#Exports the data to the excel sheet
while(b<numberofteam):
    ws.cell(row=a,column=1).value=team_numberer[b]
    ws.cell(row=a,column=2).value=Scores[b]
    ws.cell(row=a,column=3).value=Defense[b]
    a=a+1
    b=b+1

workbook.save(path)





