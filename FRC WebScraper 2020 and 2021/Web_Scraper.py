# General Note: I do not like for loops at all which is why I(Ayush Garg) only used while loops
#This means that any time you see a singular letter it is just used for the counting the iterations
#of the loop

import requests #Used to get HTML
import numpy as np #Used to solve overdetermined simultaneous equations where negative values are fine
from scipy.optimize import nnls #Used to solve overdetermined simultaneous equations where negative values are not fine
from openpyxl import load_workbook #Used to export data to excel sheet

#This method gets the stuff from the websites
#You will have to update this method every year as the stuff you are collecting changes
def get_stuff(URL, m,just_teams_array,just_auto_line_array,just_scores_array,just_climb_array, just_fouls_arrays,\
              auto_lower_array,auto_outer_array,auto_inner_array,teleop_lower_array,teleop_outer_array,\
              teleop_inner_array,control_panel_points,generator_level_points):
                #You have to put in all of the arrays you use in the method up here and then return them later so that
                # You can keep on adding on to the same arrays

    page = requests.get(URL)
    code = str(page.text) #These two lines get the HTML and put it into a string so that it can be searched


    #General Note: The 2 things you will be doing with the HTML code string
    #1. Using the find function:
        #This function finds a certain string within a larger string
        #The first input is the string the second is the starting index the final is the ending index
        #The staring and ending index will specify the range where the function will search
    #2. Indexing the string
        #This will just give you the string from the starting index to the ending index
        # The staring index is inclusive but the ending index is not inclusive
        #The syntax is string[starting index:ending index]


    #General Note:There are 2 ways to add something to an array(Technically they are called lists but who cares(not me))
        #1. Append
                # This just adds something to the end of the list
        #2. Setting a value to certain index
                #you can set a certain index equal to something if and only if that index is only defined as existing
                #or currently has a value
                #You can do this with append or when you create the array


    # Final General Note:
        #These lines:
            # numeric_filter = filter(str.isdigit, string you want to filter)
            #string you want to filter = "".join(numeric_filter)
        #Take a string and remove all non-numbers from it


    #Gets teams from TBA
    i = 1
    index = 0
    while i < 7:
        index = code.find("data-team", index + 1)
        start = index + 14
        end = index + 18
        Team = code[start:end]
        numeric_filter = filter(str.isdigit, Team)
        Team = "".join(numeric_filter)
        just_teams_array.append(int(Team))
        i = i + 1

    #Gets Red Score from TBA
    index = code.find('dScore" colspan="2"><')
    start = index + 23
    end = index + 26
    Score = code[start:end]
    numeric_filter = filter(str.isdigit, Score)
    Score = "".join(numeric_filter)
    just_scores_array.append(int(Score))

    # Gets Blue Score from TBA
    index = code.find('eScore" colspan="2"><')
    start = index + 23
    end = index + 26
    Score = code[start:end]
    numeric_filter = filter(str.isdigit, Score)
    Score = "".join(numeric_filter)
    just_scores_array.append(int(Score))

    #Gets initiation line scores
    index = code.find("Autonomous")
    start = index + 114
    end = index + 116
    a = 1
    while a < 7:
        Auto_line = code[start:end]
        if Auto_line == "ok":
            intAuto_line = 5
        if Auto_line == "re":
            intAuto_line = 0
        just_auto_line_array.append(intAuto_line)
        index = code.find("-", end)
        start = index + 1
        end = index + 3
        a = a + 1

    #Gets auto_lower_points for red team
    index=code.find("Autonomous")
    index=code.find("svg",index)
    index=code.find("svg",index+1)
    start=int(index+14)
    end=int(index+16)
    port= code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port =int( "".join(numeric_filter))
    auto_lower_array.append(int(port*2))

    # Gets auto_outer_points for red team
    index=code.find("svg",index+1)
    index=code.find("svg",index+1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    auto_outer_array.append(int(port*4))

    # Gets auto_inner_points for red team
    index = code.find("svg", index+1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port =int( "".join(numeric_filter))
    auto_inner_array.append(int(port*6 ))

    # Gets auto_lower_points for blue team
    index = code.find("svg", index+1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    auto_lower_array.append(int(port*2))

    # Gets auto_outer_points for blue team
    index = code.find("svg", index+1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    auto_outer_array.append(int(port*4))

    # Gets auto_inner_points for blue team
    index = code.find("svg", index+1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    auto_inner_array.append(int(port*6))

    # Gets teleop_lower_points for red team
    index = code.find("svg", index + 1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    teleop_lower_array.append(int(port))

    # Gets teleop_outer_points for red team
    index = code.find("svg", index + 1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    teleop_outer_array.append(int(port*2))

    # Gets teleop_inner_points for red team
    index = code.find("svg", index + 1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    teleop_inner_array.append(int(port*3))

    # Gets teleop_lower_points for blue team
    index = code.find("svg", index + 1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    teleop_lower_array.append(int(port))

    # Gets teleop_outer_points for blue team
    index = code.find("svg", index + 1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    teleop_outer_array.append(int(port*2))

    # Gets teleop_inner_points for blue team
    index = code.find("svg", index + 1)
    index = code.find("svg", index + 1)
    start = index + 14
    end = index + 16
    port = code[start:end]
    numeric_filter = filter(str.isdigit, port)
    port = int("".join(numeric_filter))
    teleop_inner_array.append(int(port*3))

    # Gets control panel points from TBA
    index=code.find("Control Panel",index)
    panel=code[index-25:index-22]
    numeric_filter = filter(str.isdigit, panel)
    panel = int("".join(numeric_filter))
    control_panel_points.append(panel)
    panel=code[index+74:index+77]
    numeric_filter = filter(str.isdigit, panel)
    panel = int("".join(numeric_filter))
    control_panel_points.append(panel)

    # Gets fouls from TBA
    index = code.find("Foul")
    foul = int(code[index + 73:index + 75])
    index = code.find("/", index)
    t_foul = code[index + 3:index + 5]
    numeric_filter = filter(str.isdigit, t_foul)
    t_foul = "".join(numeric_filter)
    foul = foul + int(t_foul)
    just_fouls_arrays.append(foul)

    # Gets tech fouls from TBA
    index = code.find("Tech Fouls")
    foul = int(code[index + 52:index + 54])
    index = code.find("/", index + 20)
    t_foul = code[index + 3:index + 5]
    numeric_filter = filter(str.isdigit, t_foul)
    t_foul = "".join(numeric_filter)
    foul = foul + int(t_foul)
    just_fouls_arrays.append(foul)

    #Gets all climb points
    k = 0
    while (k < 6):
        just_climb_array.append(0)
        k = k + 1

    index = code.find("Robot 1")
    index = code.find("title", index + 1)
    index = code.find('"', index + 9)
    if (code[index + 2] == "P"):
        just_climb_array[6 * m] = 5
    if (code[index + 2] == "H"):
        just_climb_array[6 * m] = 25

    index = code.find("title", index)
    index = code.find('"', index + 9)
    if (code[index + 2] == "P"):
        just_climb_array[6 * m + 3] = 5
    if (code[index + 2] == "H"):
        just_climb_array[6 * m + 3] = 25

    index = code.find("title", index)
    index = code.find('"', index + 9)
    if (code[index + 2] == "P"):
        just_climb_array[6 * m + 1] = 5
    if (code[index + 2] == "H"):
        just_climb_array[6 * m + 1] = 25

    index = code.find("title", index)
    index = code.find('"', index + 9)
    if (code[index + 2] == "P"):
        just_climb_array[6 * m + 4] = 5
    if (code[index + 2] == "H"):
        just_climb_array[6 * m + 4] = 25

    index = code.find("title", index)
    index = code.find('"', index + 9)
    if (code[index + 2] == "P"):
        just_climb_array[6 * m + 2] = 5
    if (code[index + 2] == "H"):
        just_climb_array[6 * m + 2] = 25


    index = code.find("title", index)
    index = code.find('"', index + 9)
    if (code[index + 2] == "P"):
        just_climb_array[6 * m + 5] = 5
    if (code[index + 2] == "H"):
        just_climb_array[6 * m + 5] = 25

    #Gets Generator level points for red team
    index=code.find("Level",index)
    index=code.find("glyphicon",index-100)
    if(code[index+20:index+21]=="o"):
        generator_level_points.append(15)
    else:
        generator_level_points.append(0)

    # Gets Generator level points for blue team
    index = code.find("glyphicon", index)
    if (code[index + 20:index + 21] == "o"):
        generator_level_points.append(15)
    else:
        generator_level_points.append(0)


    # Returns arrays so that they can updated the next time this runs through
    return just_teams_array,just_scores_array,just_auto_line_array,just_climb_array, just_fouls_arrays,auto_lower_array\
    ,auto_outer_array,auto_inner_array,teleop_lower_array,teleop_outer_array,teleop_inner_array,control_panel_points\
    ,generator_level_points


#This method takes all of the individual values of score of things like endgame and sums them up for every team then
#divides it by the number of times the team played
#You have to input an array with all of the teams in simplified format(1,4,3,2) and in the order they appeared in the
# tournament
#Along with this you have to input an array with the values that were scored for each team in the order they appeared
# in the tournament
def Averager(just_teams_array,thing_being_averaged,output,numberofteam):
    r = 1
    while r < numberofteam + 1:
        number_of_matches_per_team_array.append(just_teams_array.count(r))
        output.append(0)
        r = r + 1

    #Sums it all up
    c = 0
    while (c < numberofmatch * 6):
        x=just_teams_array[c]
        output[x-1]=output[x-1]+ thing_being_averaged[c]
        c = c + 1

    #divides it
    a=0
    while(a < numberofteam):
        output[a]=output[a]/number_of_matches_per_team_array[a]
        a=a+1

    return output

# everything above defines methods and import statements
################################################################################################

#These are most of the arrays that are used
#Their names are pretty self explanatory
#If still confused print the confusing array and analyze it to figure it out
just_teams_array = []
just_scores_array=[]
just_auto_line_array=[]
just_reduced_scores_array=[]
auto_line_per_team_array=[]
auto_line_per_match_array=[]
climb_per_match_array=[]
number_of_matches_per_team_array=[]
just_fouls_arrays=[]
just_climb_array=[]
climb_per_team_array=[]
expected_score_array=[]
defense_team_array=[]
Total_worth_array= []
playoff_matches_array=[None] * 100
auto_lower_array=[]
auto_outer_array=[]
auto_inner_array=[]
teleop_lower_array=[]
teleop_outer_array=[]
teleop_inner_array=[]
control_panel_points=[]
generator_level_points=[]


year=2020
#Where you type in the tournament ID
event_code= "mimcc"

l=1
big_loop_parameter=0
URL = "https://www.thebluealliance.com/event/"+str(year) + event_code
page = requests.get(URL)
code= str(page.text)

#Gets the number of matches from TBA
index=0
while(l<100000):
    index = code.find("2020" + event_code + "_qm"+str(l),index+1)
    if(index==-1):
        break
    l=l+1

print(l)
big_loop_parameter=l
l=1

#These next 3 arrays gather the number and type of playoff matches
#Each match gets 3 entries into the playoff array
# The first tells type of match, second tells match number, third tells round
# So 3,1,2 tells you 2nd round of first match in finals


#Gets Quarter finals 3 number IDs
a=1
j=1
c=0
while(c<1000000):
    index=code.find("qf" + str(j)+ "m" + str(a),index+1)
    if(index!=-1):
        playoff_matches_array[c]=1
        playoff_matches_array[c+1]=j
        playoff_matches_array[c+2]=a
        a=a+1
        c=c+3
    else:
        if(a!=1):
            a=1
            j=j+1
        else:
            break;

#Gets Semi finals 3 number IDs
j=1
a=1
while(c<1000000):
    index=code.find("sf" + str(j)+ "m" + str(a),index+1)
    if(index!=-1):
        playoff_matches_array[c]=2
        playoff_matches_array[c+1]=j
        playoff_matches_array[c+2]=a
        a=a+1
        c=c+3
    else:
        if(a!=1):
            a=1
            j=j+1
        else:
            break;

#Gets finals 3 number IDs
j=1
a=1
while(c<1000000):
    index=code.find("_f" + str(j)+ "m" + str(a),index+1)
    if(index!=-1):
        playoff_matches_array[c]=3
        playoff_matches_array[c+1]=j
        playoff_matches_array[c+2]=a
        a=a+1
        c=c+3
    else:
        if(a!=1):
            a=1
            j=j+1
        else:
            break;

#Gets data from qualification matches
m=0
URL = "https://www.thebluealliance.com/match/2020" +event_code +"_qm"
while l<big_loop_parameter:

    #This looks like a mess but is actually just everything that you returned in the get stuff method
    #Being set equal to its outside of method counterpart and then getting recycled into the get stuff method
    # But of course with a different match number being analyzed
    just_teams_array, just_scores_array, just_auto_line_array, just_climb_array, just_fouls_arrays,auto_lower_array,\
    auto_outer_array,auto_inner_array,teleop_lower_array,teleop_outer_array,teleop_inner_array,control_panel_points,\
    generator_level_points= get_stuff(URL+str(l) ,m,just_teams_array,just_auto_line_array,just_scores_array,\
    just_climb_array,just_fouls_arrays,auto_lower_array,auto_outer_array,auto_inner_array,teleop_lower_array,\
    teleop_outer_array,teleop_inner_array,control_panel_points,generator_level_points)

    print(l)
    m=m+1
    l=l+1

#Does the exact same thing as loop above but with playoff matches
#The if statements use the playoff matches array to help construct the URL that will be searched in the get stuff method
b=0
URL= "https://www.thebluealliance.com/match/2020" +event_code +"_"
while playoff_matches_array[b]!=None:
    match_type_adder=""
    if(playoff_matches_array[b]==1):
        match_type_adder="qf"
    if (playoff_matches_array[b] == 2):
        match_type_adder = "sf"
    if (playoff_matches_array[b] == 3):
        match_type_adder = "f"
    round_adder = playoff_matches_array[b+1]
    match_adder = playoff_matches_array[b+2]
    just_teams_array, just_scores_array, just_auto_line_array, just_climb_array, just_fouls_arrays,auto_lower_array,\
    auto_outer_array,auto_inner_array,teleop_lower_array,teleop_outer_array,teleop_inner_array,control_panel_points,\
    generator_level_points=get_stuff(URL +str(match_type_adder) + str(round_adder) + "m" + str(match_adder),m,
    just_teams_array,just_auto_line_array,just_scores_array,just_climb_array, just_fouls_arrays,auto_lower_array,\
    auto_outer_array,auto_inner_array,teleop_lower_array,teleop_outer_array,teleop_inner_array,control_panel_points,\
    generator_level_points)
    b=b+3
    m=m+1


# everything above gets the data from TBA
################################################################################################


lenght_score=len(just_scores_array)
team_numberer=set(just_teams_array) #Gets rid of duplicates
team_numberer=sorted(team_numberer) #Makes it go in ascending order

print("Teams:")
print(team_numberer)

#Simplifys the just teams array into smaller numbers
#So the smallest team number would become 1 and so on
e=1
d=0

while (e - 1 < int(len(team_numberer))):
    while (int(len(just_teams_array)) > d):
        if (team_numberer[e - 1] == just_teams_array[d]):
            just_teams_array[d] = e
        d = d + 1
    e = e + 1
    d = 0



numberofteam=int(len(team_numberer))
numberofmatch=int(lenght_score/2)

#Creates a 2-D array which has the number of teams as one dimension and the number of scores as the other
Team = [[0 for k in range(numberofteam)] for j in range(numberofmatch*2)]

#Averages the auto_line_array and climb_array with respect to team
auto_line_per_team_array=Averager(just_teams_array,just_auto_line_array,auto_line_per_team_array,numberofteam)
climb_per_team_array=Averager(just_teams_array,just_climb_array,climb_per_team_array,numberofteam)

#Puts one in the teams slot whenever it appears on the field
a=0
b=0
while(a<numberofmatch*2):
    Team[a][just_teams_array[b]-1]=1
    b=b+1
    Team[a][just_teams_array[b]-1]=1
    b=b+1
    Team[a][just_teams_array[b]-1]=1
    b=b+1
    a=a+1


# everything above gets the data into the proper format
################################################################################################

#The following lines of code use nnls which solves overdetermined simultaneous equations with least square error
#over the positive numbers. The first input is the array names team which has 0's every where except where the
#team appears in the tournament. The second is the score array.

A=np.array(Team)
B=np.array(auto_lower_array)
Auto_lower_points= nnls(A, B)[0]


A=np.array(Team)
B=np.array(auto_outer_array)
Auto_outer_points= nnls(A, B)[0]


A=np.array(Team)
B=np.array(auto_inner_array)
Auto_inner_points= nnls(A, B)[0]


A=np.array(Team)
B=np.array(teleop_lower_array)
Teleop_lower_points= nnls(A, B)[0]


A=np.array(Team)
B=np.array(teleop_outer_array)
Teleop_outer_points= nnls(A, B)[0]


A=np.array(Team)
B=np.array(teleop_inner_array)
Teleop_inner_points= nnls(A, B)[0]


A=np.array(Team)
B=np.array(control_panel_points)
Control_panel_points_calculated= nnls(A, B)[0]

A=np.array(Team)
B=np.array(generator_level_points)
Generator_level_points_calculated= nnls(A, B)[0]

#Sums up total Offense
Offense=Auto_lower_points+Auto_outer_points+Auto_inner_points+Teleop_lower_points+Teleop_outer_points+\
        Teleop_inner_points + climb_per_team_array+auto_line_per_team_array+ Control_panel_points_calculated\
        +Generator_level_points_calculated

#Creates the expected score per alliance per match which is the sum of the 3 teams offense
y=0
v=0
expected_score=0
while(y<numberofmatch*2):
    while(v<numberofteam):
        if(Team[y][v]==1):
            expected_score=expected_score+Offense[v]
        v=v+1
    y=y+1
    v=0
    expected_score_array.append(expected_score)
    expected_score=0


defense_team_array=Team
#Switches up the arrays so that the opposing alliance array is in the slot where the other alliance was
j=0
while(j<numberofmatch*2):
    defense_team_array[j],defense_team_array[j+1]=defense_team_array[j+1],defense_team_array[j]
    j=j+2

#Calculates fouls
A=np.array(defense_team_array)
B=np.array(just_fouls_arrays)
Fouls= nnls(A, B)[0]
print("Fouls:")
print(Fouls)

#Calculates Defense
#This may or may not be an accurate measure as it is very hard to prove in either direction so for now it is in beta
A=np.array(defense_team_array)
B=np.array(np.array(expected_score_array)-np.array(np.array(just_scores_array)-np.array(just_fouls_arrays)))
Defense= np.linalg.lstsq(A, B,rcond=None)[0]
print("Defense:")
print(Defense)


Total_worth_array= Offense+Defense-Fouls
print("Total_worth:")
print(Total_worth_array)

#Prints the data into the terminal
print("Broken Up by Team:")
print("Offense, Fouls, Defense, Total Worth")
f=0
while(f<numberofteam):
    print(team_numberer[f] , ":" , ", ", Offense[f], ", ", Fouls[f], ", ", Defense[f], ", ", Total_worth_array[f]  )
    f=f+1

#Is the path of the excel sheet where the data will go
path='/Users/ayushgarg/Desktop/Web Scraper Test.xlsx'
workbook = load_workbook(path)
ws = workbook.get_sheet_by_name("Sheet1")

#Finds the first unoccupied row of the excel sheet
a=1
while(2>1):
    if(ws.cell(row=a,column=1).value==None):
        break;
    a=a+1
b=0

#Exports the data to the excel sheet
while(b<numberofteam):
    ws.cell(row=a,column=1).value=team_numberer[b]
    ws.cell(row=a, column=2).value=Offense[b]
    ws.cell(row=a,column=3).value=Fouls[b]
    ws.cell(row=a, column=4).value = Defense[b]
    ws.cell(row=a, column=5).value = Total_worth_array[b]
    ws.cell(row=a, column=6).value = climb_per_team_array[b]
    ws.cell(row=a, column=7).value = auto_line_per_team_array[b]
    ws.cell(row=a, column=8).value = Auto_lower_points[b]
    ws.cell(row=a, column=9).value = Auto_outer_points[b]
    ws.cell(row=a, column=10).value = Auto_inner_points[b]
    ws.cell(row=a, column=11).value = Teleop_lower_points[b]
    ws.cell(row=a, column=12).value = Teleop_outer_points[b]
    ws.cell(row=a, column=13).value = Teleop_inner_points[b]
    ws.cell(row=a, column=14).value = Control_panel_points_calculated[b]
    ws.cell(row=a, column=15).value = generator_level_points[b]
    a=a+1
    b=b+1

workbook.save(path)



# everything above does the computations and exports it to excel along with printing it
################################################################################################


