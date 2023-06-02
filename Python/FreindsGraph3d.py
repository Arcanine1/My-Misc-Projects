from openpyxl import load_workbook #Used to export data to excel sheet
import numpy as np #Used to solve overdetermined simultaneous equations where negative values are fine
from sklearn.manifold import MDS
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

print(10)
path='/Users/ayushgarg/Desktop/Graph.xlsx'
workbook = load_workbook(path)
ws = workbook.get_sheet_by_name("Prom")

size = 8
A = np.zeros((size,size))
for i in range(1,1+size):
   for j in range(1,1+size):
      A[i-1,j-1]= ws.cell(row = i, column = j).value



mds= MDS(2, dissimilarity='precomputed')
results = mds.fit(A)

newcordinates = results.embedding_

print (newcordinates)

#labels = ["Ayush", "Kavya", "Aryan", "Nandan", "Saket", "Ankur", "Ansh", "Caroline", "Charlie", "Jaan", "Jasper", "Mattew", "Yohaan", "Zahra" ]
labels = ["Aryan", "Gary", "Sohan", "VC", "Saket", "Kavya", "Prisha", "Nethra"]



fig, ax = plt.subplots()
ax.scatter(newcordinates[:,0],newcordinates[:,1])
for i,label in enumerate(labels):
    ax.annotate(label,newcordinates[i])

plt.show()


# fig = plt.figure()
# ax = fig.add_subplot(111, projection='3d')
# ax.scatter(newcordinates[:,0],newcordinates[:,1],newcordinates[:,2])
# for i,label in enumerate(labels):
#     ax.text(newcordinates[i,0],newcordinates[i,1],newcordinates[i,2],label)
# plt.show()